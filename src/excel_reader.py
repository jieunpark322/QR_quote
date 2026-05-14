"""작성된 엑셀 견적서를 읽어 QuoteDocument로 변환."""
from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from .excel_template import (
    CELLS,
    ITEM_END_ROW,
    ITEM_START_ROW,
    SHEET_QUOTE,
)
from .models import Counterparty, LineItem, QuoteDocument, Totals


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip()[:10])
        except ValueError:
            return None
    return None


def _to_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _load_catalog_by_name(project_root: Path) -> dict[str, dict]:
    path = project_root / "catalog" / "products.json"
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return {p["name"]: p for p in data.get("products", [])}


def _slugify_counterparty(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\"<>|\s]+", "", name)
    return cleaned[:8] if cleaned else "고객"


def read_quote_from_excel(xlsx_path: Path, project_root: Path) -> QuoteDocument:
    if not xlsx_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {xlsx_path}")

    catalog = _load_catalog_by_name(project_root)
    wb = load_workbook(str(xlsx_path), data_only=True)
    if SHEET_QUOTE not in wb.sheetnames:
        raise ValueError(f"'{SHEET_QUOTE}' 시트를 찾을 수 없습니다. 템플릿이 맞는지 확인해주세요.")
    ws = wb[SHEET_QUOTE]

    brand_id = _to_str(ws[CELLS["brand_id"]].value) or "softment"
    document_id = _to_str(ws[CELLS["document_id"]].value)
    subject = _to_str(ws[CELLS["subject"]].value) or "(건명 미입력)"
    issued = _to_date(ws[CELLS["issued_date"]].value) or date.today()
    valid_until = _to_date(ws[CELLS["valid_until"]].value)

    cp = Counterparty(
        name=_to_str(ws[CELLS["cp_name"]].value) or "(수신처 미입력)",
        registration_number=_to_str(ws[CELLS["cp_reg_no"]].value),
        address=_to_str(ws[CELLS["cp_address"]].value),
        contact_name=_to_str(ws[CELLS["cp_contact_name"]].value),
        contact_title=_to_str(ws[CELLS["cp_contact_title"]].value),
        email=_to_str(ws[CELLS["cp_email"]].value),
    )

    items: list[LineItem] = []
    for row in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        name = _to_str(ws.cell(row=row, column=1).value)
        if not name:
            continue

        raw_desc = ws.cell(row=row, column=2).value
        raw_qty = ws.cell(row=row, column=3).value
        raw_period = ws.cell(row=row, column=4).value
        raw_price = ws.cell(row=row, column=5).value
        raw_notes = ws.cell(row=row, column=7).value

        desc = _to_str(raw_desc)
        # 수식 미계산 시 카탈로그에서 채우기
        catalog_entry = catalog.get(name)
        if catalog_entry:
            if not desc:
                desc = catalog_entry.get("description") or None
            if raw_price is None or raw_price == 0:
                raw_price = catalog_entry.get("unit_price", 0)

        try:
            qty = float(raw_qty) if raw_qty not in (None, "") else None
        except (TypeError, ValueError):
            qty = None
        try:
            period = float(raw_period) if raw_period not in (None, "") else None
        except (TypeError, ValueError):
            period = None
        try:
            unit_price = float(raw_price) if raw_price not in (None, "") else 0.0
        except (TypeError, ValueError):
            unit_price = 0.0

        if unit_price == 0 and not qty and not period:
            continue

        items.append(LineItem(
            name=name,
            description=desc,
            qty=qty,
            period=period,
            unit_price=unit_price,
            currency="KRW",
            notes=_to_str(raw_notes),
        ))

    notes = _to_str(ws[CELLS["notes"]].value)

    # subtotal만 계산해서 넘기고, vat/vat_rate/total은 renderer에서 ensure_totals가 채움
    subtotal = sum(i.amount for i in items)

    if not document_id:
        document_id = f"Q-{issued.strftime('%Y%m%d')}-{_slugify_counterparty(cp.name)}"

    return QuoteDocument(
        document_id=document_id,
        document_type="quote",
        brand_id=brand_id,
        issued_date=issued,
        valid_until=valid_until,
        counterparty=cp,
        subject=subject,
        line_items=items,
        totals=Totals(subtotal=subtotal, currency="KRW"),
        clauses=[],
        notes=notes,
    )
