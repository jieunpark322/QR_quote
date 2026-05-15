"""견적서 입력용 엑셀 템플릿 생성기.

드롭다운으로 품목을 선택하면 VLOOKUP으로 설명/단가가 자동 채워지고,
드롭다운에 없는 품목은 직접 입력할 수 있습니다.
"""
from __future__ import annotations

import json
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .labels import DocumentLabels, load_labels


SHEET_QUOTE = "견적서"
SHEET_CATALOG = "품목카탈로그"

ITEM_START_ROW = 21
ITEM_END_ROW = 40
TOTALS_START_ROW = 43

# 셀 좌표 (리더가 동일하게 참조)
CELLS = {
    "brand_id": "B6",
    "document_id": "B7",
    "subject": "B8",
    "issued_date": "B9",
    "valid_until": "B10",
    "cp_name": "B13",
    "cp_reg_no": "B14",
    "cp_address": "B15",
    "cp_contact_name": "B16",
    "cp_contact_title": "D16",
    "cp_email": "B17",
    "notes": "A48",
}


def _load_products(catalog_path: Path) -> list[dict]:
    data = json.loads(catalog_path.read_text(encoding="utf-8"))
    return data["products"]


def _set_column_widths(ws) -> None:
    # A:항목, B:설명, C:단가, D:기간(횟수), E:수량, F:공급가, G:비고
    widths = {"A": 22, "B": 28, "C": 14, "D": 10, "E": 8, "F": 16, "G": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _label_cell(ws, addr: str, text: str) -> None:
    c = ws[addr]
    c.value = text
    c.font = Font(name="맑은 고딕", size=10, bold=True, color="333333")
    c.fill = PatternFill("solid", fgColor="F0F2F8")
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    c.border = _thin_border()


def _input_cell(ws, addr: str, placeholder: str = "") -> None:
    c = ws[addr]
    c.value = placeholder
    c.font = Font(name="맑은 고딕", size=10)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    c.border = _thin_border()


def _section_header(ws, row: int, text: str, color: str = "0B3D91") -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    c = ws.cell(row=row, column=1)
    c.value = text
    c.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[row].height = 24


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _build_catalog_sheet(wb: Workbook, products: list[dict]):
    ws = wb.create_sheet(SHEET_CATALOG)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20

    notice = ws.cell(row=1, column=1)
    ws.merge_cells("A1:E1")
    notice.value = ("ℹ 이 시트는 드롭다운과 단가 자동 채우기에 사용됩니다. "
                    "영구적인 변경은 catalog/products.json 파일을 편집하세요.")
    notice.font = Font(name="맑은 고딕", size=9, italic=True, color="888888")
    notice.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 22

    headers = ["품목명", "설명", "단가", "통화", "코드"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col)
        c.value = h
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0B3D91")
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.border = _thin_border()
    ws.row_dimensions[2].height = 22

    for idx, p in enumerate(products, start=3):
        row_data = [p["name"], p.get("description", ""), p["unit_price"],
                    p.get("currency", "KRW"), p["code"]]
        for col, value in enumerate(row_data, start=1):
            c = ws.cell(row=idx, column=col)
            c.value = value
            c.font = Font(name="맑은 고딕", size=10)
            c.alignment = Alignment(vertical="center",
                                    horizontal="right" if col == 3 else "left",
                                    indent=0 if col == 3 else 1)
            if col == 3:
                c.number_format = '#,##0'
            c.border = _thin_border()

    ws.freeze_panes = "A3"
    return len(products)


def _build_quote_sheet(wb: Workbook, products: list[dict],
                       brand_id: str, default_issued: date,
                       default_valid_until: date,
                       labels: DocumentLabels) -> None:
    ws = wb.create_sheet(SHEET_QUOTE, 0)
    _set_column_widths(ws)

    # 제목
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "📋 견적서 작성 폼"
    title.font = Font(name="맑은 고딕", size=18, bold=True, color="0B3D91")
    title.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 32

    # 안내
    ws.merge_cells("A3:G3")
    guide = ws["A3"]
    guide.value = ("📌 사용법:  품목을 드롭다운에서 고르면 설명·단가가 자동 입력됩니다.  "
                   "수량 × 기간(회) × 단가 = 금액으로 자동 계산됩니다.  "
                   "각 행의 비고 칸에 특이사항을 자유롭게 적을 수 있습니다.")
    guide.font = Font(name="맑은 고딕", size=9, italic=True, color="555555")
    guide.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
    ws.row_dimensions[3].height = 32

    # ■ 발행 정보
    _section_header(ws, 5, "■ 발행 정보")
    _label_cell(ws, "A6", "브랜드 ID")
    _input_cell(ws, CELLS["brand_id"], brand_id)
    _label_cell(ws, "A7", "문서번호 (선택, 비우면 자동)")
    _input_cell(ws, CELLS["document_id"], "")
    _label_cell(ws, "A8", "건명")
    ws.merge_cells("B8:G8")
    _input_cell(ws, CELLS["subject"], "")
    _label_cell(ws, "A9", "발행일")
    c = ws[CELLS["issued_date"]]
    c.value = default_issued
    c.number_format = "yyyy-mm-dd"
    c.font = Font(name="맑은 고딕", size=10)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    c.border = _thin_border()
    _label_cell(ws, "A10", "유효기간")
    c = ws[CELLS["valid_until"]]
    c.value = default_valid_until
    c.number_format = "yyyy-mm-dd"
    c.font = Font(name="맑은 고딕", size=10)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    c.border = _thin_border()

    # ■ 수신처 정보
    _section_header(ws, 12, "■ 수신처 (고객) 정보")
    _label_cell(ws, "A13", "회사명")
    ws.merge_cells("B13:G13")
    _input_cell(ws, CELLS["cp_name"], "")
    _label_cell(ws, "A14", "사업자등록번호")
    _input_cell(ws, CELLS["cp_reg_no"], "")
    _label_cell(ws, "A15", "주소")
    ws.merge_cells("B15:G15")
    _input_cell(ws, CELLS["cp_address"], "")
    _label_cell(ws, "A16", "담당자")
    _input_cell(ws, CELLS["cp_contact_name"], "")
    _label_cell(ws, "C16", "직책")
    _input_cell(ws, CELLS["cp_contact_title"], "")
    _label_cell(ws, "A17", "Email")
    ws.merge_cells("B17:G17")
    _input_cell(ws, CELLS["cp_email"], "")

    # ■ 품목 내역
    _section_header(ws, 19, "■ 품목 내역  (드롭다운 선택 또는 직접 입력 모두 가능)")

    th = labels.quote.table_headers
    # 컬럼 순서: 항목·설명·단가·기간(횟수)·수량·공급가·비고
    headers = [th.name, th.description, th.unit_price, th.period, th.qty, th.amount, th.notes]
    header_row = 20
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col)
        c.value = h
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="0B3D91")
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.border = _thin_border()
    ws.row_dimensions[header_row].height = 22

    catalog_last_row = 2 + len(products)
    dropdown_range = f"{SHEET_CATALOG}!$A$3:$A${catalog_last_row}"
    catalog_lookup = f"{SHEET_CATALOG}!$A$3:$C${catalog_last_row}"

    dv = DataValidation(
        type="list",
        formula1=f"={dropdown_range}",
        allow_blank=True,
        showErrorMessage=False,
        showDropDown=False,
    )
    dv.prompt = "드롭다운에서 선택하거나 직접 입력하세요"
    dv.promptTitle = "품목"
    dv.showInputMessage = True
    dv_target = f"A{ITEM_START_ROW}:A{ITEM_END_ROW}"
    ws.add_data_validation(dv)
    dv.add(dv_target)

    for row in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        # A: 항목 (드롭다운)
        c = ws.cell(row=row, column=1)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        c.border = _thin_border()

        # B: 설명 (VLOOKUP col 2, 수정 가능)
        c = ws.cell(row=row, column=2)
        c.value = f'=IFERROR(VLOOKUP(A{row},{catalog_lookup},2,FALSE),"")'
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        c.border = _thin_border()

        # C: 단가 (VLOOKUP col 3, 수정 가능)
        c = ws.cell(row=row, column=3)
        c.value = f'=IFERROR(VLOOKUP(A{row},{catalog_lookup},3,FALSE),0)'
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="right", indent=1)
        c.border = _thin_border()
        c.number_format = '#,##0'

        # D: 기간(횟수)
        c = ws.cell(row=row, column=4)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.border = _thin_border()
        c.number_format = '#,##0'

        # E: 수량
        c = ws.cell(row=row, column=5)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.border = _thin_border()
        c.number_format = '#,##0'

        # F: 공급가 (= [수량 or 1] × [기간 or 1] × 단가). 단가가 비면 빈 칸.
        c = ws.cell(row=row, column=6)
        c.value = (f'=IF(C{row}="","",'
                   f'IF(D{row}="",1,D{row})*IF(E{row}="",1,E{row})*C{row})')
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="right", indent=1)
        c.border = _thin_border()
        c.number_format = '#,##0'
        c.fill = PatternFill("solid", fgColor="FAFBFD")

        # G: 비고 (자유 입력)
        c = ws.cell(row=row, column=7)
        c.font = Font(name="맑은 고딕", size=10)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        c.border = _thin_border()

    # ■ 합계
    _section_header(ws, TOTALS_START_ROW - 1, "■ 합계  (자동 계산)")

    ql = labels.quote.labels
    vat_rate = labels.quote.vat_rate
    totals = [
        (ql.subtotal,
         f"=SUM(F{ITEM_START_ROW}:F{ITEM_END_ROW})", False),
        (f"{ql.vat} ({int(vat_rate * 100)}%)",
         f"=ROUND(G{TOTALS_START_ROW}*{vat_rate},0)", False),
        (ql.total,
         f"=G{TOTALS_START_ROW}+G{TOTALS_START_ROW+1}", True),
    ]
    for offset, (label, formula, highlight) in enumerate(totals):
        row = TOTALS_START_ROW + offset
        lc = ws.cell(row=row, column=6)
        lc.value = label
        lc.font = Font(name="맑은 고딕", size=11,
                       bold=True,
                       color="FFFFFF" if highlight else "333333")
        lc.fill = PatternFill("solid", fgColor="0B3D91" if highlight else "F0F2F8")
        lc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
        lc.border = _thin_border()

        vc = ws.cell(row=row, column=7)
        vc.value = formula
        vc.font = Font(name="맑은 고딕", size=12 if highlight else 11,
                       bold=True,
                       color="FFFFFF" if highlight else "111111")
        vc.fill = PatternFill("solid", fgColor="0B3D91" if highlight else "FAFBFD")
        vc.alignment = Alignment(vertical="center", horizontal="right", indent=1)
        vc.border = _thin_border()
        vc.number_format = '"₩"#,##0'
        ws.row_dimensions[row].height = 22 if highlight else 20

    # ■ 비고 (문서 전체)
    notes_section_row = TOTALS_START_ROW + 4
    _section_header(ws, notes_section_row, "■ 비고  (문서 전체 - 선택)")
    notes_row = notes_section_row + 1
    ws.merge_cells(start_row=notes_row, start_column=1,
                   end_row=notes_row, end_column=7)
    c = ws.cell(row=notes_row, column=1)
    c.value = ""
    c.font = Font(name="맑은 고딕", size=10)
    c.alignment = Alignment(vertical="top", horizontal="left", indent=1, wrap_text=True)
    c.border = _thin_border()
    ws.row_dimensions[notes_row].height = 60

    # 행 높이 조정
    for r in [5, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17, 19, 20]:
        if r not in (5, 12, 19):
            ws.row_dimensions[r].height = 22

    # 데이터 행 약간 크게
    for r in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        ws.row_dimensions[r].height = 22

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A21"


def build_template(project_root: Path, output_path: Path,
                   brand_id: str = "softment",
                   issued_date: date | None = None,
                   valid_days: int | None = None) -> Path:
    """엑셀 입력 템플릿을 생성합니다.

    valid_days 미지정 시 config/labels.json의 default_validity_days 사용.
    """
    catalog_path = project_root / "catalog" / "products.json"
    if not catalog_path.exists():
        raise FileNotFoundError(f"품목 카탈로그를 찾을 수 없습니다: {catalog_path}")
    products = _load_products(catalog_path)

    labels = load_labels(project_root)
    if valid_days is None:
        valid_days = labels.quote.default_validity_days

    issued_date = issued_date or date.today()
    valid_until = issued_date + timedelta(days=valid_days)

    wb = Workbook()
    # 기본 시트 제거
    default_sheet = wb.active
    wb.remove(default_sheet)

    _build_quote_sheet(wb, products, brand_id, issued_date, valid_until, labels)
    _build_catalog_sheet(wb, products)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
